import os
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OEM_FILE = os.path.join(BASE_DIR, "OEM.xlsx")


# -------- OEM EXCEL DATA --------
def read_oem_data():

    workbook = load_workbook(OEM_FILE)
    sheet = workbook.active

    protocol_data = {}

    for row in sheet.iter_rows(values_only=True):

        if not row or not row[0]:
            continue  # skip empty rows

        protocol = str(row[0]).strip()
        supported = str(row[1]).strip() if row[1] else ""
        details = str(row[2]).strip() if row[2] else ""

        protocol_data[protocol] = {
            "supported": supported,
            "details": details
        }

    return protocol_data


# -------- MAIN FUNCTION --------
def run_OEM_test():

    result_data = {
        "oem_protocol_data": read_oem_data()
    }

    return result_data